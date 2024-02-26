import pandas as pd 
import os
from openpyxl import  load_workbook
"""
função recebe um dataframe e transformando em um arquivo excel
"""

def converter_dataframe(df:pd.DataFrame,caminho:str,filename:str,template:str): #-> str:df.to_excel(f'{caminho}/{filename}.xlsx',index=False)
     # Carrega o template
    if not os.path.exists(caminho):
        os.makedirs("data/output")
    else:
        wb = load_workbook(template)

     # Seleciona a planilha ativa
    ws = wb.active

    # Escreve o DataFrame na planilha
    #necessario definir no for de colunas um enumerador  para indice receber o valor do enumerdor e col o nome da coluna,
    # o row precisa ser +2 para ignorar o cabeçalho do template
    for row in df.index:
        for indice, col in enumerate(df.columns):
            ws.cell(row=row+2, column = indice+ 1).value = df.loc[row, col]

    # Salva o arquivo
    caminho_completo = f'{caminho}/{filename}.xlsx'
    wb.save(caminho_completo)


    
   
    
    
    
   
   


    
        